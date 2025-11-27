#!/usr/bin/env python
# coding: utf-8

# In[25]:


#Verileri aldık ayrı kategoriledik sonrasında none içeren verilerin tümünü silerek filtreledik 




# In[27]:


from openpyxl import load_workbook, Workbook

# 1. Adım: Mevcut 'nessus.xlsx' dosyasını yükleyin
file_path = 'nessus.xlsx'  # Dosya yolu
wb_old = load_workbook(file_path)
ws_old = wb_old.active

# 2. Adım: Yeni bir workbook (Excel dosyası) oluşturun
wb_new = Workbook()
ws_new = wb_new.active

# 3. Adım: Yeni dosyaya başlıkları ekleyin
headers = ['Plugin ID', 'Risk', 'Host', 'Protocol', 'Port', 'Name']
for col_num, header in enumerate(headers, 1):
    ws_new.cell(row=1, column=col_num, value=header)

# 4. Adım: Eski dosyadaki verileri satır satır okuyun ve yeni dosyaya ayırarak yazın
for row_num, row in enumerate(ws_old.iter_rows(min_row=2, values_only=True), start=2):
    if row:  # Satır boş değilse
        # Satırdaki verileri virgülle ayırarak uygun sütunlara yerleştirin
        values = row[0].split(',')  # Her bir satırı virgülle ayırın

        # Verileri uygun başlıkların altına ekleyin
        ws_new.cell(row=row_num, column=1, value=values[0])  # Plugin ID -> A sütunu
        ws_new.cell(row=row_num, column=2, value=values[1].strip('"'))  # Risk -> B sütunu
        ws_new.cell(row=row_num, column=3, value=values[2].strip('"'))  # Host -> C sütunu
        ws_new.cell(row=row_num, column=4, value=values[3].strip('"'))  # Protocol -> D sütunu
        ws_new.cell(row=row_num, column=5, value=values[4].strip('"'))  # Port -> E sütunu
        ws_new.cell(row=row_num, column=6, value=values[5].strip('"'))  # Name -> F sütunu

# 5. Adım: Yeni düzenlenmiş dosyayı kaydedin
new_file_path = 'nessus.new.xlsx'
wb_new.save(new_file_path)

print(f"Veriler başarıyla '{new_file_path}' dosyasına kaydedildi.")


# In[28]:


from openpyxl import load_workbook, Workbook

# 1. Adım: Mevcut 'nessus.xlsx' dosyasını yükleyin
file_path = 'nessus.new.xlsx'  # Dosya yolu
wb_old = load_workbook(file_path)
ws_old = wb_old.active

# 2. Adım: Yeni bir workbook (Excel dosyası) oluşturun
wb_new = Workbook()
ws_new = wb_new.active

# 3. Adım: Yeni dosyaya başlıkları ekleyin
headers = ['Plugin ID', 'Risk', 'Host', 'Protocol', 'Port', 'Name']
for col_num, header in enumerate(headers, 1):
    ws_new.cell(row=1, column=col_num, value=header)

# 4. Adım: Eski dosyadaki verileri satır satır okuyun ve risk değeri 'None' olmayanları yeni dosyaya ekleyin
new_row_num = 2  # Yeni dosyada verilerin yazılacağı satır numarası
for row in ws_old.iter_rows(min_row=2, values_only=True):
    if row and row[1] != 'None':  # Risk değeri 'None' değilse
        # Verileri uygun sütunlara yerleştirin
        for col_num, value in enumerate(row, 1):
            ws_new.cell(row=new_row_num, column=col_num, value=value)
        new_row_num += 1  # Sonraki satıra geç

# 5. Adım: Yeni düzenlenmiş dosyayı kaydedin
new_file_path = 'filtered_nessus.xlsx'
wb_new.save(new_file_path)

print(f"Risk değeri 'None' olan veriler silindi ve '{new_file_path}' dosyasına kaydedildi.")


# In[29]:


import os
import re
from openpyxl import load_workbook

def sanitize_filename(name):
    # Windows'ta geçersiz karakterleri temizleme
    return re.sub(r'[<>:"/\\|?*\x00-\x1F]', '', name)

def organize_files_from_excel(input_excel):
    try:
        # Excel dosyasını oku
        print(f"'{input_excel}' dosyasından veriler okunuyor ve klasörler oluşturuluyor...")
        wb = load_workbook(input_excel)
        ws = wb.active
        
        # Klasör ve dosya oluşturma işlemleri
        name_dict = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            plugin_id, risk, host, protocol, port, name = row
            if name not in name_dict:
                name_dict[name] = []
            name_dict[name].append(f"{host}:{port}")
        
        for name, ip_port_list in name_dict.items():
            # Geçerli dizin adı ve dosya yolu
            sanitized_name = sanitize_filename(name)
            dir_name = os.path.join('output', sanitized_name)
            if not os.path.exists(dir_name):
                os.makedirs(dir_name)
            
            # data.txt dosyasının yolu
            file_path = os.path.join(dir_name, 'data.txt')
            
            # IP:port formatında veri ekle
            with open(file_path, mode='w') as txt_file:
                for ip_port in ip_port_list:
                    txt_file.write(f"{ip_port}\n")
        
        print("Klasörler ve dosyalar başarıyla oluşturuldu.")
    
    except Exception as e:
        print(f"Klasör oluşturma veya dosya yazma sırasında bir hata oluştu: {e}")

if __name__ == "__main__":
    # Dosya yolunu buraya yerleştirin
    input_excel_dosya = r'filtered_nessus.xlsx'  # Excel dosyasının yolu
    
    # Excel dosyasını işle ve klasörler oluştur
    organize_files_from_excel(input_excel_dosya)
    
    print("Veriler işleme alındı ve dosyalar oluşturuldu.")



# In[ ]:





# In[ ]:




